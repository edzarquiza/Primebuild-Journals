import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile
import io
import base64
from datetime import datetime
import re

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Primebuild Payroll Journals",
    page_icon="📓",
    layout="wide",
)

# ── Brand theme ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=Space+Grotesk:wght@400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
    background-color: #050505;
    color: #f0f0f0;
}
.stApp { background-color: #050505; }

h1, h2, h3 {
    font-family: 'Space Grotesk', sans-serif;
    background: linear-gradient(90deg, #00c4b4, #0066ff);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}

.block-container { padding-top: 2rem; }

.stFileUploader > div {
    background: #0f0f0f;
    border: 1px solid #1a1a2e;
    border-radius: 12px;
}
.stFileUploader label { color: #f0f0f0 !important; }

.stDateInput > div > div > input {
    background: #0f0f0f !important;
    color: #f0f0f0 !important;
    border: 1px solid #1a1a2e !important;
    border-radius: 8px !important;
}
.stDateInput label { color: #f0f0f0 !important; }

div[data-testid="stButton"] > button {
    background: linear-gradient(90deg, #00c4b4, #0066ff);
    color: white;
    font-family: 'Space Grotesk', sans-serif;
    font-weight: 600;
    border: none;
    border-radius: 8px;
    padding: 0.6rem 2rem;
    font-size: 1rem;
    transition: opacity 0.2s;
    width: 100%;
}
div[data-testid="stButton"] > button:hover { opacity: 0.88; }

div[data-testid="stDownloadButton"] > button {
    background: linear-gradient(90deg, #00c4b4, #0066ff);
    color: white;
    font-family: 'Space Grotesk', sans-serif;
    font-weight: 600;
    border: none;
    border-radius: 8px;
    padding: 0.6rem 2rem;
    font-size: 1rem;
    width: 100%;
}

.metric-card {
    background: #0f0f0f;
    border: 1px solid #1a1a2e;
    border-radius: 12px;
    padding: 1.2rem 1.5rem;
    margin-bottom: 0.5rem;
}
.metric-label {
    font-size: 0.75rem;
    color: #888;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 0.3rem;
}
.metric-value {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 1.6rem;
    font-weight: 600;
    background: linear-gradient(90deg, #00c4b4, #0066ff);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}

.file-tag {
    display: inline-block;
    background: #0f0f0f;
    border: 1px solid #00c4b4;
    border-radius: 6px;
    padding: 0.2rem 0.7rem;
    font-size: 0.8rem;
    color: #00c4b4;
    margin: 0.2rem;
}

.success-box {
    background: #0a1a0f;
    border: 1px solid #00c4b4;
    border-radius: 10px;
    padding: 1rem 1.5rem;
    margin-top: 1rem;
}
.error-box {
    background: #1a0a0a;
    border: 1px solid #ff4444;
    border-radius: 10px;
    padding: 1rem 1.5rem;
    margin-top: 0.5rem;
    color: #ff9999;
}
.divider {
    height: 1px;
    background: linear-gradient(90deg, #00c4b4, #0066ff);
    margin: 1.5rem 0;
    opacity: 0.3;
}
</style>
""", unsafe_allow_html=True)

# ── Logo ──────────────────────────────────────────────────────────────────────
def get_logo_b64():
    try:
        with open("logo.jpg", "rb") as f:
            return base64.b64encode(f.read()).decode()
    except FileNotFoundError:
        return None

logo_b64 = get_logo_b64()

col_logo, col_title = st.columns([1, 5])
with col_logo:
    if logo_b64:
        st.markdown(f'<img src="data:image/jpeg;base64,{logo_b64}" style="width:80px;margin-top:6px;">', unsafe_allow_html=True)
with col_title:
    st.markdown("## Payroll Journals Automation")
    st.markdown('<p style="color:#888;margin-top:-0.5rem;font-size:0.9rem;">Generate GL journal download files from KeyPay raw exports</p>', unsafe_allow_html=True)

st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
STATE_CWI = {'NSW': 10, 'QLD': 40, 'VIC': 20, 'ROL': 11, 'SVS': 85, 'CON': 50}

LOOKUP_K = {
    'CBA Cheque Account', 'PAYG Withholding Tax', 'Payroll Tax Payable',
    'Superannuation Clearing', 'Payroll Clearing', 'Other Payroll Deductions',
    'Annual Leave Entitlements', 'Sick & Personal Leave Entitlem',
    'RDO Accrual', 'LSL Provision', 'Operations SLPHUni Accrual',
    'Provision - WCompensation'
}

OUTPUT_HEADERS = [
    'Dissection', 'Description', 'Costing Work Id', 'Job', 'Cost Code',
    'Cost Type', 'Resource Code', 'GL Account', 'Quantity', 'UOM', 'Amount',
    'Normal Value', 'Allowance Value', 'On Cost Value', 'Tax Code',
    'Tax Percentage', 'Tax Amount', 'Internal Reference', 'External Reference',
    'Asset Work Id', 'Asset', 'Small Order Revenue', 'Text',
]

# ── Core transformation logic ─────────────────────────────────────────────────
def parse_filename(filename):
    """Extract state, frequency, default CWI from raw filename.
    Handles WComp in any capitalisation (WCOMP, WComp, Wcomp, wcomp).
    """
    name = filename.replace('.xlsx', '').replace('.xlsm', '')
    state = name[:3].upper()
    # Case-insensitive WComp detection anywhere in the filename
    is_wcomp = 'WCOMP' in name.upper()
    freq = 'WC' if is_wcomp else name[4:6].upper()
    cwi = STATE_CWI.get(state, 10)
    return state, freq, cwi


def process_raw_file(file_bytes, filename, payment_date_str):
    """
    Transform one raw JNL export into a list of output rows,
    replicating the VBA macro logic exactly.
    """
    state, freq, default_cwi = parse_filename(filename)

    # Format payment date
    dt = datetime.strptime(payment_date_str, '%d/%m/%Y')
    fmt_date = dt.strftime('%d/%m/%Y')
    time_date_fname = dt.strftime('%Y%m%d')

    # Internal reference string
    # WComp files still use FN in the internal ref (filename contains _FN_)
    upper_name = filename.upper().replace('.XLSX','')
    if 'WCOMP' in upper_name:
        m = re.search(r'WCOMP_([A-Z]+)_', upper_name)
        raw_freq = m.group(1) if m else 'FN'
    else:
        raw_freq = upper_name[4:6]
    internal_ref = f"{state} {raw_freq} PAY {fmt_date}"

    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name='Journal', header=None)
    data = df.iloc[1:].reset_index(drop=True)  # skip header row

    rows = []
    dissection = 0

    for _, row in data.iterrows():
        acct_no   = str(int(row[3])) if pd.notna(row[3]) else ''
        acct_name = str(row[4]) if pd.notna(row[4]) else ''
        desc      = str(row[5]) if pd.notna(row[5]) else ''
        amount    = row[6] if pd.notna(row[6]) else 0
        tax_code  = str(row[10]) if pd.notna(row[10]) and str(row[10]) != 'nan' else ''
        keypay    = str(row[11]) if pd.notna(row[11]) and str(row[11]) != 'nan' else ''
        dim2      = str(row[13]) if pd.notna(row[13]) and str(row[13]) != 'nan' else ''

        dissection += 1

        # Col B: Description
        col_b = f"{desc} {acct_name}".strip()

        # Col C: Costing Work Id (initial derivation from dim2 formula)
        in_lookup = acct_name in LOOKUP_K
        n_char4 = dim2[3] if len(dim2) > 3 else ''
        if not in_lookup and n_char4 in ('C', 'D'):
            col_c = int(dim2[:2])
        else:
            col_c = default_cwi

        # Defaults
        col_d = ''   # Job
        col_e = ''   # Cost Code
        col_f = ''   # Cost Type

        # Parse dim2 if it contains slashes (len >= 10)
        if len(dim2) >= 10 and '/' in dim2:
            parts = dim2.split('/')
            cw_id    = parts[0] if len(parts) > 0 else ''
            job_code = parts[1] if len(parts) > 1 else ''
            cost_code = parts[2] if len(parts) > 2 else ''
            cost_type = parts[3] if len(parts) > 3 else ''

            if cost_type == 'RV':
                cost_type = 'RC'

            if cw_id.isdigit():
                col_c = int(cw_id)
            col_d = job_code
            col_e = cost_code
            col_f = cost_type

            # ROL special cost type rules
            if state == 'ROL':
                if len(cost_code) > 3 and cost_code[3] == '-':
                    col_f = 'RC'
                elif len(cost_code) > 7:
                    col_f = cost_type
                elif len(cost_code) == 5:
                    col_f = 'LB'
                else:
                    col_f = 'RC'
                if job_code and job_code[0] == 'D':
                    col_f = 'LB'

            # R-starting job code -> CA
            if job_code and job_code[0] == 'R':
                col_f = 'CA'
        else:
            # No valid dim2 -> clear Job and Cost Type
            col_d = ''
            col_f = ''

        # Col H: GL Account
        # Rows with a parsed job code (dim2 with slashes) get empty GL Account
        # Rows without (lookup lines, Prime Build overrides) get CWI+AccountNo
        has_job = len(dim2) >= 10 and '/' in dim2
        col_h = '' if has_job else f"{default_cwi}{acct_no}"

        # Area code override for "Prime Build" rows (not HOLDINGS)
        if keypay.lower().startswith('prime build') and 'HOLDINGS' not in keypay:
            right6 = keypay[-6:]
            col_h = right6
            col_c = int(right6[:2]) if right6[:2].isdigit() else col_c
            col_d = ''
            col_e = ''
            col_f = ''

        # NSW_WK rollouts override
        if state == 'NSW' and freq == 'WK':
            if 'rollouts' in col_b.lower() or '- ROL' in col_b:
                col_c = 11

        rows.append({
            'Dissection': dissection,
            'Description': col_b,
            'Costing Work Id': col_c,
            'Job': col_d if col_d else None,
            'Cost Code': col_e,
            'Cost Type': col_f if col_f else None,
            'Resource Code': '',
            'GL Account': col_h,
            'Quantity': None,
            'UOM': None,
            'Amount': amount,
            'Normal Value': None,
            'Allowance Value': None,
            'On Cost Value': None,
            'Tax Code': tax_code if tax_code else None,
            'Tax Percentage': None,
            'Tax Amount': None,
            'Internal Reference': internal_ref,
            'External Reference': str(int(float(row[1]))) if pd.notna(row[1]) else '',
            'Asset Work Id': None,
            'Asset': None,
            'Small Order Revenue': None,
            'Text': col_b,
        })

    # Determine output filename
    if freq == 'WC':
        out_name = f"WComp {state} {time_date_fname} JNL.xlsx"
    else:
        out_name = f"{state} {freq} {time_date_fname} JNL.xlsx"

    return rows, out_name, state, freq, default_cwi


def build_output_workbook(rows, state, freq, cwi, payment_date_str):
    """Build the JRNDOWNLD-format Excel workbook from processed rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "JRNDOWNLD"

    dt = datetime.strptime(payment_date_str, '%d/%m/%Y')

    # ── Styles matching actual output exactly ──
    # Title: bold size 20, no fill
    title_font    = Font(name='Arial', bold=True, size=20)
    # Batch header labels (row 2): small grey
    batch_font    = Font(name='Arial', size=10)
    # Column headers (row 4): bold white on dark navy #002856
    hdr_font      = Font(name='Arial', bold=True, size=11, color='FFFFFFFF')
    navy_fill     = PatternFill('solid', fgColor='002856')
    # Data rows: salmon/pink #E5B8B7 (theme colour 5 + tint 0.6)
    data_font     = Font(name='Arial', size=11)
    row_fill      = PatternFill('solid', fgColor='E5B8B7')

    # ── Row 1: Title ──
    ws['A1'] = 'General Ledger Journal Download'
    ws['A1'].font = title_font
    ws.merge_cells('A1:Y1')

    # ── Row 2: Batch header labels ──
    batch_labels = ['Batch', 'Work Id', 'Date', 'Period', 'Batch Type',
                    'Debit Check Sum', 'Line Count Check', 'Text']
    for col_i, lbl in enumerate(batch_labels, start=1):
        cell = ws.cell(row=2, column=col_i, value=lbl)
        cell.font = batch_font

    # ── Row 3: Batch data ──
    ws.cell(row=3, column=2, value=cwi).font = batch_font
    date_cell = ws.cell(row=3, column=3, value=dt)
    date_cell.number_format = 'DD/MM/YYYY'
    date_cell.font = batch_font
    ws.cell(row=3, column=5, value='G').font = batch_font

    # ── Row 4: Column headers (navy bg, white bold text) ──
    col_headers = OUTPUT_HEADERS + ['Error Message', 'POSTED']
    for col_i, hdr in enumerate(col_headers, start=1):
        cell = ws.cell(row=4, column=col_i, value=hdr)
        cell.font = hdr_font
        cell.fill = navy_fill

    # ── Row 5: Empty with POSTED marker ──
    ws.cell(row=5, column=25, value='POSTED').font = data_font

    # ── Data rows (start row 6) with salmon fill on all columns A-Y ──
    total_cols = len(col_headers)  # 25 (A-Y)
    for row_data in rows:
        r = ws.max_row + 1
        # Apply fill across all columns first
        for col_i in range(1, total_cols + 1):
            ws.cell(row=r, column=col_i).fill = row_fill
        # Then write data
        for col_i, key in enumerate(OUTPUT_HEADERS, start=1):
            val = row_data.get(key)
            cell = ws.cell(row=r, column=col_i, value=val)
            cell.font = data_font
            cell.fill = row_fill
            if key in ('GL Account', 'Cost Code', 'External Reference'):
                cell.number_format = '@'
            if key == 'Amount' and val is not None:
                cell.number_format = '#,##0.00'

    # ── Column widths ──
    col_widths = {
        'A': 10, 'B': 50, 'C': 16, 'D': 12, 'E': 14, 'F': 10,
        'G': 14, 'H': 14, 'I': 10, 'J': 8,  'K': 14, 'L': 14,
        'M': 16, 'N': 14, 'O': 10, 'P': 14, 'Q': 12, 'R': 24,
        'S': 18, 'T': 14, 'U': 10, 'V': 18, 'W': 50,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = 'A6'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── UI: Inputs ────────────────────────────────────────────────────────────────
col_left, col_right = st.columns([3, 2])

with col_left:
    st.markdown("### 📂 Upload Raw Journal Files")
    uploaded_files = st.file_uploader(
        "Upload one or more raw JNL export files",
        type=['xlsx'],
        accept_multiple_files=True,
        help="Files should follow the naming convention: STATE_FREQ_YYYYMMDD_JNL_Raw.xlsx"
    )

with col_right:
    st.markdown("### 📅 Payment Date")
    payment_date = st.date_input(
        "Enter the payment date",
        value=None,
        format="DD/MM/YYYY",
        help="This date will appear in the Internal Reference of each journal line."
    )
    st.markdown('<p style="font-size:0.8rem;color:#888;">Format: DD/MM/YYYY</p>', unsafe_allow_html=True)

    if uploaded_files:
        st.markdown('<div style="margin-top:1rem;"></div>', unsafe_allow_html=True)
        st.markdown("**Files detected:**")
        for f in uploaded_files:
            state, freq, cwi = parse_filename(f.name)
            st.markdown(f'<span class="file-tag">{state} {freq} (CWI: {cwi})</span>', unsafe_allow_html=True)

# ── Metrics row ──────────────────────────────────────────────────────────────
if uploaded_files:
    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
    m1, m2, m3 = st.columns(3)
    with m1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Files Uploaded</div>
            <div class="metric-value">{len(uploaded_files)}</div>
        </div>""", unsafe_allow_html=True)
    with m2:
        total_rows = 0
        for f in uploaded_files:
            try:
                df = pd.read_excel(io.BytesIO(f.read()), sheet_name='Journal', header=None)
                total_rows += len(df) - 1
                f.seek(0)
            except Exception:
                f.seek(0)
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Total Journal Lines</div>
            <div class="metric-value">{total_rows:,}</div>
        </div>""", unsafe_allow_html=True)
    with m3:
        pay_str = payment_date.strftime('%d/%m/%Y') if payment_date else '—'
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Payment Date</div>
            <div class="metric-value" style="font-size:1.2rem;">{pay_str}</div>
        </div>""", unsafe_allow_html=True)

# ── Generate button ───────────────────────────────────────────────────────────
st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

if st.button("⚙️ Generate Journal Files", disabled=not (uploaded_files and payment_date)):
    payment_date_str = payment_date.strftime('%d/%m/%Y')

    errors = []
    output_files = {}

    progress = st.progress(0, text="Processing files...")
    total = len(uploaded_files)

    for i, uploaded_file in enumerate(uploaded_files):
        try:
            file_bytes = uploaded_file.read()
            rows, out_name, state, freq, cwi = process_raw_file(
                file_bytes, uploaded_file.name, payment_date_str
            )
            xlsx_bytes = build_output_workbook(rows, state, freq, cwi, payment_date_str)
            output_files[out_name] = xlsx_bytes
            progress.progress((i + 1) / total, text=f"Processed: {uploaded_file.name}")
        except Exception as e:
            errors.append(f"{uploaded_file.name}: {str(e)}")
            progress.progress((i + 1) / total, text=f"Error on: {uploaded_file.name}")

    progress.empty()

    if errors:
        for err in errors:
            st.markdown(f'<div class="error-box">⚠️ {err}</div>', unsafe_allow_html=True)

    if output_files:
        # Build ZIP
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fname, fbytes in output_files.items():
                zf.writestr(fname, fbytes)
        zip_buf.seek(0)

        zip_name = f"Payroll_Journals_{payment_date.strftime('%Y%m%d')}.zip"

        st.markdown(f"""
        <div class="success-box">
            ✅ <strong>{len(output_files)} journal file(s) generated successfully.</strong><br>
            <span style="font-size:0.85rem;color:#aaa;">Payment date: {payment_date_str} &nbsp;|&nbsp; Files: {', '.join(output_files.keys())}</span>
        </div>""", unsafe_allow_html=True)

        st.download_button(
            label=f"⬇️ Download {zip_name}",
            data=zip_buf.getvalue(),
            file_name=zip_name,
            mime="application/zip",
        )

elif not uploaded_files:
    st.markdown('<p style="color:#555;text-align:center;padding:1rem;">Upload raw journal files above to get started.</p>', unsafe_allow_html=True)
elif not payment_date:
    st.markdown('<p style="color:#555;text-align:center;padding:1rem;">Select a payment date to continue.</p>', unsafe_allow_html=True)

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
st.markdown('<p style="color:#333;font-size:0.75rem;text-align:center;">Dexterous · Primebuild Payroll Journals Automation</p>', unsafe_allow_html=True)
